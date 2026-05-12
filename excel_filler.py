from openpyxl import load_workbook


def fill_excel(data):
    print(data)

    #load excel template
    wb = load_workbook("Template/solar_emplate.xlsx")

    ws = wb.active

    # Example cells
    ws["H20"] = data.get("units")#units here
    ws["I20"] = data.get("amount")

    output_path = "outputs/final_output.xlsx"

    wb.save(output_path)

    return output_path
